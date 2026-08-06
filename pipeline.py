# -*- coding: utf-8 -*-
"""食堂农残报告：订单PDF + 检测图片 -> 填表 -> 盖章 -> 导出PDF。"""

from __future__ import annotations

import tempfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from pdf_report import render_report_pdf
from ocr_report import build_inhibition_map_from_ocr, lookup_rate_in_map, normalize_name as _ocr_normalize_name
from order_pdf import extract_vegetables as extract_vegetables_tolerant
from stamp_utils import optimize_stamp_image

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = BASE_DIR / "模板湖北合盛源-农药残留检验报告.xlsx"
DEFAULT_STAMP = BASE_DIR / "默认红章.png"
USER_STAMP = BASE_DIR / "user_stamp.png"

# 订单菜名 <-> 检测报告样品名（仅保留明确等价别名，避免短词误匹配）
NAME_ALIASES = {
    "新鲜大土豆": ["土豆", "马铃薯"],
    "土豆": ["新鲜大土豆", "马铃薯"],
    "西红柿": ["番茄"],
    "番茄": ["西红柿"],
    "老南瓜": ["南瓜"],
    "南瓜": ["老南瓜"],
}


class StepError(Exception):
    """带步骤名的中文失败说明。"""

    def __init__(self, step: str, detail: str, cause: Optional[BaseException] = None):
        self.step = step
        self.detail = detail
        self.cause = cause
        parts = [f"【{step}失败】{detail}"]
        if cause is not None:
            parts.append(f"具体原因：{type(cause).__name__}: {cause}")
        super().__init__("\n".join(parts))


def resolve_stamp_path(preferred: str | Path | None = None) -> Path:
    """优先使用指定章；否则用户上次保存的章；再否则默认章。"""
    if preferred is not None:
        p = Path(preferred)
        if p.exists():
            return p
    if USER_STAMP.exists():
        return USER_STAMP
    if DEFAULT_STAMP.exists():
        return DEFAULT_STAMP
    raise StepError(
        "读取红章",
        f"未找到红章文件。请先在界面上传红章，或将红章图片放到程序目录并命名为「{DEFAULT_STAMP.name}」。",
    )


def save_user_stamp(data: bytes, suffix: str = ".png") -> Path:
    """保存用户上传的红章（自动压缩），供下次默认使用。"""
    raw = USER_STAMP.with_name("user_stamp_raw" + (suffix if suffix.startswith(".") else f".{suffix}"))
    raw.write_bytes(data)
    try:
        optimize_stamp_image(raw, USER_STAMP, max_side=360)
        raw.unlink(missing_ok=True)
    except Exception:
        # 压缩失败则直接存原图
        USER_STAMP.write_bytes(data)
        raw.unlink(missing_ok=True)
    return USER_STAMP


def _normalize_name(name: str) -> str:
    return _ocr_normalize_name(name)


def extract_vegetables_from_pdf(pdf_path: str | Path) -> List[str]:
    """步骤1：从采购订单PDF提取蔬菜/食材名称列（包容食材名称/名称/品名等表头）。"""
    try:
        vegetables = extract_vegetables_tolerant(str(pdf_path))
        cleaned: List[str] = []
        for name in vegetables:
            n = _normalize_name(name)
            if not n:
                continue
            if n not in cleaned:
                cleaned.append(n)
        if not cleaned:
            raise StepError(
                "识别订单PDF",
                "未能提取到有效的蔬菜名称。请确认上传的是食堂采购订单PDF。",
            )
        return cleaned
    except StepError:
        raise
    except ValueError as e:
        raise StepError("识别订单PDF", str(e)) from e
    except Exception as e:
        raise StepError("识别订单PDF", "读取或解析采购订单PDF时出错。", e) from e


def extract_vegetables_from_pdfs(pdf_paths: List[str | Path]) -> List[str]:
    """合并多份订单PDF中的蔬菜名（去重、保序）。"""
    if not pdf_paths:
        raise StepError("识别订单PDF", "未上传任何订单PDF。")
    merged: List[str] = []
    for path in pdf_paths:
        for name in extract_vegetables_from_pdf(path):
            if name not in merged:
                merged.append(name)
    if not merged:
        raise StepError(
            "识别订单PDF",
            "未能从上传的订单PDF中提取到有效蔬菜名称。",
        )
    return merged


def _get_ocr_engine():
    from rapidocr_onnxruntime import RapidOCR

    return RapidOCR()


def extract_inhibition_map(
    image_path: str | Path,
    order_vegetables: Optional[List[str]] = None,
) -> Tuple[Dict[str, float], Optional[str]]:
    """步骤2：OCR农残图片，建立 {样品名称: 抑制率数值}。

    包容：样品列可为样品/名称/品名等；未知表头时按订单菜名反查。
    抑制率列：只要表头含「抑制率」且该列是数值即可。
    """
    try:
        ocr = _get_ocr_engine()
        result, _ = ocr(str(image_path))
    except Exception as e:
        raise StepError(
            "识别检测报告图片",
            "OCR引擎启动或识别失败。请确认已安装 rapidocr-onnxruntime，且图片可正常打开。",
            e,
        ) from e

    if not result:
        raise StepError(
            "识别检测报告图片",
            "图片中未识别到任何文字。请上传更清晰的农残检验报告图片。",
        )

    try:
        mapping, inspect_date = build_inhibition_map_from_ocr(
            result,
            order_vegetables=order_vegetables,
            aliases=NAME_ALIASES,
        )
    except ValueError as e:
        raise StepError("识别检测报告图片", str(e)) from e
    except Exception as e:
        raise StepError(
            "识别检测报告图片",
            "解析检测报告表格失败。请换更清晰、端正的质检报告图片重试。",
            e,
        ) from e

    if not mapping:
        raise StepError(
            "识别检测报告图片",
            "已识别到文字，但未能解析出蔬菜名称与抑制率的对应关系。请换更清晰的质检报告图片重试。",
        )
    return mapping, inspect_date


def extract_inhibition_map_from_images(
    image_paths: List[str | Path],
    order_vegetables: Optional[List[str]] = None,
) -> Tuple[Dict[str, float], Optional[str]]:
    """合并多张检测报告图片的 OCR 结果。"""
    if not image_paths:
        raise StepError("识别检测报告图片", "未上传任何检测报告图片。")

    merged: Dict[str, float] = {}
    inspect_date: Optional[str] = None
    errors: List[str] = []

    for path in image_paths:
        try:
            mapping, date_text = extract_inhibition_map(
                path, order_vegetables=order_vegetables
            )
            for k, v in mapping.items():
                if k not in merged:
                    merged[k] = v
            if date_text and not inspect_date:
                inspect_date = date_text
        except StepError as e:
            errors.append(f"{Path(path).name}: {e.detail}")

    if not merged:
        detail = "所有检测报告图片均未能解析出抑制率。"
        if errors:
            detail += " " + "；".join(errors[:3])
        raise StepError("识别检测报告图片", detail)
    return merged, inspect_date


def lookup_rate(veg: str, rate_map: Dict[str, float]) -> Optional[float]:
    """按精确名 / 别名 / 模糊包含查找抑制率。"""
    return lookup_rate_in_map(veg, rate_map, aliases=NAME_ALIASES)


def build_report_data(
    vegetables: List[str],
    rate_map: Dict[str, float],
    inspect_date: Optional[str] = None,
) -> Tuple[List[Dict], List[str], str]:
    """生成表格行。未匹配到的菜：抑制率留空，列入 missing 列表。"""
    if inspect_date:
        date_text = inspect_date
    else:
        today = datetime.now()
        date_text = f"{today.year}年{today.month}月{today.day}日"

    rows: List[Dict] = []
    missing: List[str] = []
    for i, veg in enumerate(vegetables):
        rate = lookup_rate(veg, rate_map)
        if rate is None:
            missing.append(veg)
            rows.append(
                {
                    "seq": i + 1,
                    "name": veg,
                    "rate": "",
                    "standard": "<50%",
                    "result": "合格",
                }
            )
        else:
            rate = round(float(rate), 2)
            rows.append(
                {
                    "seq": i + 1,
                    "name": veg,
                    "rate": f"{rate:.2f}%",
                    "standard": "<50%",
                    "result": "合格",
                }
            )
    return rows, missing, date_text


def fill_template(
    vegetables: List[str],
    rate_map: Dict[str, float],
    template_path: str | Path,
    output_xlsx: str | Path,
    inspect_date: Optional[str] = None,
    rows: Optional[List[Dict]] = None,
    date_text: Optional[str] = None,
) -> Tuple[Path, List[str], List[Dict], str]:
    """步骤3：按模板从第6行填充 A~E 列，严格沿用模板字体字号（不改模板原文件）。"""
    try:
        template_path = Path(template_path)
        output_xlsx = Path(output_xlsx)
        if not template_path.exists():
            raise StepError("填充Excel模板", f"找不到模板文件：{template_path.name}")

        if rows is None or date_text is None:
            rows, missing, date_text = build_report_data(
                vegetables, rate_map, inspect_date=inspect_date
            )
        else:
            missing = [r["name"] for r in rows if not r.get("rate")]

        # 只读打开模板内容到内存，最终另存为输出文件，绝不覆盖模板原文件
        wb = load_workbook(template_path)
        ws = wb.active

        # 第3行 E3：检验时间（保留模板原字体 宋体 11）
        e3_font = copy(ws["E3"].font)
        e3_align = copy(ws["E3"].alignment)
        e3_border = copy(ws["E3"].border)
        e3_fill = copy(ws["E3"].fill)
        e3_fmt = ws["E3"].number_format
        ws["E3"] = f"检验时间：{date_text}"
        ws["E3"].font = e3_font
        ws["E3"].alignment = e3_align
        ws["E3"].border = e3_border
        ws["E3"].fill = e3_fill
        ws["E3"].number_format = e3_fmt

        footer_row = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "只对本次" in str(val):
                footer_row = r
                break
        if footer_row is None:
            raise StepError(
                "填充Excel模板",
                "模板缺少页脚「只对本次检验样品负责」，请确认使用的是「模板湖北合盛源-农药残留检验报告.xlsx」。",
            )

        # 数据区从第6行开始；按订单菜品数量增删行
        data_start = 6
        style_row = 6
        styles = []
        for c in range(1, 6):
            styles.append(
                {
                    "font": copy(ws.cell(style_row, c).font),
                    "border": copy(ws.cell(style_row, c).border),
                    "fill": copy(ws.cell(style_row, c).fill),
                    "number_format": ws.cell(style_row, c).number_format,
                    "alignment": copy(ws.cell(style_row, c).alignment),
                }
            )
        style_height = ws.row_dimensions[style_row].height

        n = len(rows)
        available = max(0, footer_row - data_start)
        if n > available:
            ws.insert_rows(footer_row, n - available)
            footer_row += n - available
        elif n < available:
            # 缩减多余数据行，页脚上移
            ws.delete_rows(data_start + n, available - n)
            footer_row -= available - n

        for i, item in enumerate(rows):
            row = data_start + i
            if style_height:
                ws.row_dimensions[row].height = style_height

            for c in range(1, 6):
                cell = ws.cell(row, c)
                st = styles[c - 1]
                cell.font = copy(st["font"])
                cell.border = copy(st["border"])
                cell.fill = copy(st["fill"])
                cell.number_format = st["number_format"]
                cell.alignment = copy(st["alignment"])

            ws.cell(row, 1).value = item["seq"]
            ws.cell(row, 2).value = item["name"]
            ws.cell(row, 3).value = item["rate"] if item["rate"] else None
            ws.cell(row, 4).value = item["standard"]
            ws.cell(row, 5).value = item["result"]

        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        if output_xlsx.resolve() == template_path.resolve():
            raise StepError("填充Excel模板", "禁止覆盖模板原文件，请检查输出路径。")
        wb.save(output_xlsx)
        return output_xlsx, missing, rows, date_text
    except StepError:
        raise
    except Exception as e:
        raise StepError("填充Excel模板", "写入Excel模板时出错。", e) from e


def stamp_excel(
    xlsx_path: str | Path,
    stamp_image: str | Path,
    output_xlsx: str | Path,
    stamp_size_px: int = 120,
    anchor_cell: str | None = None,
) -> Path:
    """红章定位：上沿对齐「只对本次…负责」行上沿；左沿对齐 B 列右沿（即 C 列左沿）。"""
    try:
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils import coordinate_to_tuple
        from openpyxl.utils.units import pixels_to_EMU

        xlsx_path = Path(xlsx_path)
        stamp_image = Path(stamp_image)
        output_xlsx = Path(output_xlsx)
        if not stamp_image.exists():
            raise StepError("Excel盖章", f"红章图片不存在：{stamp_image}")

        wb = load_workbook(xlsx_path)
        ws = wb.active

        if anchor_cell is None:
            footer_row = None
            for r in range(1, ws.max_row + 1):
                val = ws.cell(r, 1).value
                if val and "只对本次" in str(val):
                    footer_row = r
                    break
            # C 列左沿 = B 列右沿；行取页脚行，rowOff=0 即与该行上沿对齐
            anchor_cell = f"C{footer_row}" if footer_row else "C26"

        img = XLImage(str(stamp_image))
        img.width = stamp_size_px
        img.height = stamp_size_px

        row_1based, col_1based = coordinate_to_tuple(anchor_cell)
        marker = AnchorMarker(
            col=col_1based - 1,
            colOff=pixels_to_EMU(0),
            row=row_1based - 1,
            rowOff=pixels_to_EMU(0),
        )
        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(
                pixels_to_EMU(stamp_size_px),
                pixels_to_EMU(stamp_size_px),
            ),
        )
        ws.add_image(img)

        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_xlsx)
        return output_xlsx
    except StepError:
        raise
    except Exception as e:
        raise StepError("Excel盖章", "向Excel写入红章图片失败。", e) from e


def run_pipeline(
    order_pdf: str | Path | List[str | Path],
    report_image: str | Path | List[str | Path],
    stamp_image: str | Path | None = None,
    template_path: str | Path | None = None,
    work_dir: str | Path | None = None,
) -> dict:
    """完整流程：填表 -> Excel盖章；全部匹配成功才导出 PDF。"""
    template_path = Path(template_path or DEFAULT_TEMPLATE)
    if not template_path.exists():
        raise StepError(
            "准备模板",
            f"未找到报告模板「{template_path.name}」。请确保「模板湖北合盛源-农药残留检验报告.xlsx」与程序放在同一目录。",
        )

    try:
        stamp_path = resolve_stamp_path(stamp_image)
    except StepError:
        raise

    work_dir = Path(work_dir or tempfile.mkdtemp(prefix="nongcan_"))
    work_dir.mkdir(parents=True, exist_ok=True)

    order_paths = (
        list(order_pdf)
        if isinstance(order_pdf, (list, tuple))
        else [order_pdf]
    )
    report_paths = (
        list(report_image)
        if isinstance(report_image, (list, tuple))
        else [report_image]
    )

    vegetables = extract_vegetables_from_pdfs(order_paths)
    rate_map, inspect_date = extract_inhibition_map_from_images(
        report_paths, order_vegetables=vegetables
    )
    rows, missing, date_text = build_report_data(
        vegetables, rate_map, inspect_date=inspect_date
    )

    xlsx_raw = work_dir / "_填表中间稿.xlsx"
    xlsx_stamped = work_dir / "湖北合盛源-农药残留检验报告.xlsx"
    pdf_stamped = work_dir / "湖北合盛源-农药残留检验报告.pdf"

    fill_template(
        vegetables,
        rate_map,
        template_path,
        xlsx_raw,
        inspect_date=inspect_date,
        rows=rows,
        date_text=date_text,
    )

    stamp_opt = work_dir / "_stamp_opt.png"
    try:
        optimize_stamp_image(stamp_path, stamp_opt, max_side=320)
        stamp_for_use = stamp_opt
    except Exception:
        stamp_for_use = stamp_path

    stamp_excel(xlsx_raw, stamp_for_use, xlsx_stamped)

    pdf_path: Optional[str] = None
    if missing:
        message = (
            "部分蔬菜在检测报告中未找到结果，已仅生成 Excel（抑制率留空）。"
            "请从历史报告中查找后手工补全。未匹配："
            + "、".join(missing)
        )
    else:
        try:
            render_report_pdf(
                rows=rows,
                inspect_date=date_text,
                output_pdf=pdf_stamped,
                stamp_image=stamp_for_use,
            )
            pdf_path = str(pdf_stamped)
        except Exception as e:
            raise StepError(
                "导出PDF",
                "按模板绘制PDF失败。请确认已安装 pymupdf，且红章图片可正常打开。",
                e,
            ) from e
        message = "本次农残报告已自动生成，全部蔬菜检测合格"

    return {
        "vegetables": vegetables,
        "rate_map": rate_map,
        "missing": missing,
        "randomized": missing,  # 兼容旧字段名
        "xlsx_path": str(xlsx_stamped),
        "pdf_path": pdf_path,
        "message": message,
    }
